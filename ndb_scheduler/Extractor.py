import click
import collections
import openpyxl as xls


class Extractor:
    pretext = {
        'visits': 'INSERT INTO visits (`id`, `Date`, `Topic_id`, `Confirmed`, `Status`) VALUES ',
        'visits_fritids': 'INSERT INTO visits (`id`, `Date`, `Topic_id`, `Group_id`, `Confirmed`, `Status`) VALUES ',
        'groups': 'INSERT INTO groups (`id`, `Name`, `Segment`, `StartYear`, `NumberStudents`, `Status`, `User_id`, `School_id`) VALUES ',
        'users': 'INSERT INTO users (`id`, `FirstName`, `LastName`, `Mail`, `School_id`, `Role`, `Status`, `MessageSettings`) VALUES ',
        'colleagues': 'INSERT INTO colleagues_visits (`visit_id`, `user_id`) VALUES ',
        'foreign_key_disabled': 'SET FOREIGN_KEY_CHECKS=0;\n'
    }

    def __init__(self, workbook_file):

        self.file = workbook_file
        self.book = xls.load_workbook(filename=workbook_file)
        self.topics = None
        self.values = None
        self.staff = None
        self.existing_users = None
        self.schools = None
        self.visit_id = None
        self.user_id = None
        self.group_id = None
        self.visit_result_file = 'extracted_visits.sql'
        self.group_result_file = 'extracted_groups_users.sql'

    def extract_visits_as_sql(self, file_mode="a"):
        calendar = self.book['calendar']
        _v = self.get_values()


        text = self.pretext['foreign_key_disabled']
        for column in calendar.iter_cols(min_col=2):
            col = [c.value for c in column]
            upper_col = col[_v['first_row_staff'] - 1: _v['last_row_staff'] - 1]
            lower_col = col[_v['first_row_subtractions'] - 1: _v['last_row_subtractions'] - 1]

            occurrences = collections.Counter(upper_col)
            occurrences.subtract(dict(collections.Counter(lower_col)))
            occurrences = dict(occurrences)

            occurrences = {
                self.get_topic_id_from_letter(k): round(float(v) / float(self.get_staff_count_for_topic(k)))
                for k, v in occurrences.items()
                if k
            }

            for topic_id, count in occurrences.items():
                dstring = col[2].strftime("%Y-%m-%d")

                for _ in range(int(count)):
                    self.visit_id = 1 + self.get_visit_id()
                    text += self.pretext['visits']
                    text += f"({self.visit_id}, '{dstring}', {topic_id}, 0, 1);\n"
                    for cell in column:
                        if not (_v['first_row_staff'] < cell.row < _v['last_row_staff']):
                            continue
                        if cell.value and self.get_topic_id_from_letter(cell.value) == topic_id:
                            user_acronym = calendar.cell(row=cell.row, column=1).value.lower()
                            user_id = self.get_staff()[user_acronym]
                            text += self.pretext['colleagues']
                            text += f"({self.visit_id}, {user_id});\n"

        click.echo(text)
        with open(self.visit_result_file, file_mode) as insert_file:
            insert_file.write(text)

    def extract_fritids_as_sql(self, file_mode="a"):

        calendar_fritids = self.book['calendar_fritids']
        V = self.get_values()
        fritids_groups = self.get_fritids_groups()

        text = self.pretext['foreign_key_disabled']
        for column in calendar_fritids.iter_cols(min_col=2):
            date = column[2].value
            if date is None:
                continue

            dstring = date.strftime("%Y-%m-%d")
            valid_rows = range(V['first_row_staff'], V['last_row_staff'])
            col = [c for c in column if c.row in valid_rows and c.value is not None]
            for cell in col:
                school_list = fritids_groups[cell.value]
                nr_visits_for_school = school_list["nr_visits"]
                school_groups = school_list["groups"]
                count = school_list["count"]
                index = nr_visits_for_school % count
                group_id = school_groups[index]

                self.visit_id = 1 + self.get_visit_id()
                text += self.pretext['visits_fritids']
                text += f"({self.visit_id}, '{dstring}', {V['fritids_topic_id']}, {group_id}, 0, 1);\n"

                user_acronym = calendar_fritids.cell(row=cell.row, column=1).value
                user_id = self.get_staff()[user_acronym]

                text += self.pretext['colleagues']
                text += f"({self.visit_id}, {user_id});\n"
                fritids_groups[cell.value]["nr_visits"] += 1

        click.echo(text)
        with open(self.visit_result_file, file_mode) as insert_file:
            insert_file.write(text)

    def get_topic_id_from_letter(self, topic_letter):
        topics = self.get_topics()
        return topics[topic_letter.lower()]["topic_id"]

    def get_staff_count_for_topic(self, topic_letter):
        topics = self.get_topics()
        return topics[topic_letter.lower()]["staff_count"]

    def get_staff(self):
        if self.staff:
            return self.staff

        staff_sheet = self.book['staff']
        self.staff = {k.lower(): v for k, v in staff_sheet.iter_rows(values_only=True) if k is not None}

        return self.staff

    def get_value(self, name):
        values = self.get_values()
        return values[name]

    def get_values(self):
        if self.values:
            return self.values

        value_sheet = self.book['manual_values']
        self.values = {k: v
                       for k, v
                       in value_sheet.iter_rows(values_only=True, min_row=2, max_col=2)
                       if k is not None}

        return self.values

    def get_topics(self):
        if self.topics:
            return self.topics

        topics_sheet = self.book['topics']
        self.topics = {topic_letter.lower(): {"topic_id": topic_id, "staff_count": staff_count}
                       for topic_id, topic_letter, _, staff_count
                       in topics_sheet.iter_rows(values_only=True, max_col=4)
                       if topic_letter is not None}

        return self.topics

    def get_visit_id(self):
        if self.visit_id is not None:
            return self.visit_id
        self.visit_id = self.get_value('max_visit_id')
        return self.visit_id

    def get_user_id(self):
        if self.user_id is not None:
            return self.user_id
        self.user_id = self.get_value('max_user_id')
        return self.user_id

    def get_group_id(self):
        if self.group_id is not None:
            return self.group_id
        self.group_id = self.get_value('max_group_id')
        return self.group_id

    def get_fritids_groups(self):
        fri_groups_sheet = self.book['groups_fritids']

        fri_groups_rows = [r for r in fri_groups_sheet.iter_rows(values_only=True, max_col=2) if r[0] is not None]

        schools = set([r[0] for r in fri_groups_rows])
        fri_groups = {s: {"groups": [], "count": 0, "nr_visits": 0} for s in schools}

        for row in fri_groups_rows:
            school_id, group_id = row
            fri_groups[school_id]["groups"].append(group_id)
            fri_groups[school_id]["count"] += 1

        return fri_groups

    def get_school_id(self, long_name):
        schools = self.get_schools()
        if long_name in schools:
            return schools[long_name]
        return None

    def get_schools(self):
        if self.schools is not None:
            return self.schools

        school_sheet = self.book['skolor']
        self.schools = {k: v for v, k in school_sheet.iter_rows(min_row=2, values_only=True)}

        return self.schools

    def get_existing_users(self):
        if self.existing_users is not None:
            return self.existing_users
        user_sheet = self.book['existing_users']
        self.existing_users = {k.strip().lower(): int(v) for v, k in user_sheet.iter_rows(max_col=2, values_only=True)}

        return self.existing_users

    def sheet_exists(self, sheet_name):
        return sheet_name in self.book

    def step_a(self, segment=2):
        class_sheet = self.book['source']
        rows = [r for r in class_sheet.iter_rows(min_row=2) if str(r[1].value) == str(segment)]

        new_sheet = self.book.create_sheet('step_a')
        new_sheet.append(['Exclude?', None, None, None, 'Cut', 'Name', None])

        schools = {}

        for row in rows:
            exclude = None
            nr_students = int(row[4].value)
            segment = row[1].value
            school_id = self.get_school_id(row[0].value)
            if (school_id is not None) and (school_id not in schools):
                schools[school_id] = []
            class_name = row[2].value

            if (school_id is None) or (class_name in schools[school_id]) or (nr_students < 6):
                exclude = 'x'
            else:
                schools[school_id].append(class_name)
            teacher = row[3].value
            cut = None
            if (len(str(teacher).split()) > 3) and (exclude is None):
                cut = 1
            new_row = [exclude, school_id, segment, class_name, cut, teacher, nr_students]
            new_sheet.append(new_row)
            print(new_row)

        self.book.save(self.file)
        # os.startfile(self.file)

    def step_b(self):
        a_sheet = self.book['step_a']
        b_sheet = self.book.create_sheet('step_b')

        rows = [r for r in a_sheet.iter_rows(min_row=2) if r[0].value is None]
        b_sheet.append([None, None, 'Mail', None, None, None, None])

        for row in rows:
            row_values = [r.value for r in row]
            school, segment, class_name, cut, teacher, nr_students = row_values[1:]

            cut = 1 if cut is None else cut
            first_name, last_name = None, None
            if teacher is not None:
                first_parts = str(teacher).split(',')[0].split()
                last_name = " ".join(first_parts[:cut]).strip()
                first_name = " ".join(first_parts[cut:]).strip()
            new_row = [first_name, last_name, None, school, segment, class_name, nr_students]
            b_sheet.append(new_row)
            print(new_row)

        self.book.save(self.file)
        # os.startfile(self.file)

    def step_c(self, file_mode='a'):
        b_sheet = self.book['step_b']
        rows = b_sheet.iter_rows(min_row=2)

        start_year = self.get_value('start_year')
        text = self.pretext['foreign_key_disabled']
        added_users = {}
        for row in rows:
            row_values = [r.value for r in row]
            fname, lname, mail, school_id, segment, class_name, nr_students = row_values
            if mail is not None:
                mail = mail.lower().strip()

            if mail in self.get_existing_users():
                user_id = self.get_existing_users()[mail]
                add_user = False
            elif mail in added_users:
                user_id = added_users[mail]
                add_user = False
            elif mail is not None:
                self.user_id = 1 + self.get_user_id()
                user_id = self.user_id
                add_user = True
                added_users[mail] = user_id
            else:
                user_id = 'NULL'
                add_user = False

            nr_students = 'NULL' if nr_students is None else nr_students

            self.group_id = 1 + self.get_group_id()
            text += self.pretext['groups']
            text += f"({self.group_id}, '{class_name}', '{segment}', {start_year}, "
            text += f"{nr_students}, 1, {user_id}, {school_id});\n"
            if add_user:
                fname = 'NULL' if fname is None else fname
                lname = 'NULL' if lname is None else lname
                text += self.pretext['users']
                text += f"({user_id}, '{fname}', '{lname}', '{mail}', '{school_id}', 4, 1, 6);\n"

        print(text)
        with open(self.group_result_file, file_mode) as insert_file:
            insert_file.write(text)


if __name__ == '__main__':
    file = 'C:/Users/friheh001/Sigtuna kommun/Intern - ' \
           'Dokument/Administration/klasslistor_läsårsdata/data_från_IST_analys/test/klasslista_ak5.xlsx'
    # file = '/mnt/d/Sigtuna kommun/Intern - Dokument/Administration/klasslistor_läsårsdata/data_från_IST_analys/test' \
    #       '/klasslista_ak5.xlsx'
    ex = Extractor(file)
    ex.step_b()
